import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import copy, io, zipfile, os, re, uuid, traceback
from pathlib import Path

# ══════════════════════════════════════════════════════════════════
# 매핑
# ══════════════════════════════════════════════════════════════════
COMPETENCY_MAP = {
    "Position":     [9, 10, 17, 18, 22, 31],
    "Personality":  [5, 13, 14, 21, 24, 28],
    "Relationship": [6,  7, 23, 25, 30, 32],
    "Results":      [8, 16, 29, 33],
    "Development":  [4, 12, 20, 27],
    "Principles":   [11, 15, 19, 26],
}
SKILL_MAP = {
    "우호성":     [4, 12, 20, 27],
    "동기유발":   [5, 13, 21, 28],
    "자문":       [6, 14],
    "협력제휴":   [7, 15, 22, 29],
    "협상거래":   [8, 16, 23, 30],
    "합리적설득": [9, 17, 24, 31],
    "합법화":     [10, 18, 25, 32],
    "강요":       [11, 19, 26, 33],
}
SOFT_SKILLS = ["우호성", "동기유발", "자문"]
HARD_SKILLS = ["협력제휴", "협상거래", "합리적설득", "합법화", "강요"]
COMP_ROW  = {"Position":4,"Personality":5,"Relationship":6,
              "Results":7,"Development":8,"Principles":9}
SKILL_ROW = {"우호성":12,"동기유발":13,"자문":14,"협력제휴":15,
              "협상거래":16,"합리적설득":17,"합법화":18,"강요":19}

# ══════════════════════════════════════════════════════════════════
# 계산
# ══════════════════════════════════════════════════════════════════
def avg_rows(scores, rows):
    vals = [float(scores.get(str(r-3), 0)) for r in rows]
    return round(sum(vals)/len(vals), 2) if vals else 0.0

def compute(scores):
    c = {k: avg_rows(scores, v) for k, v in COMPETENCY_MAP.items()}
    s = {k: avg_rows(scores, v) for k, v in SKILL_MAP.items()}
    return {"competency": c, "skill_raw": s,
            "soft_avg": round(sum(s[k] for k in SOFT_SKILLS)/3, 2),
            "hard_avg": round(sum(s[k] for k in HARD_SKILLS)/5, 2)}

# ══════════════════════════════════════════════════════════════════
# 파싱
# ══════════════════════════════════════════════════════════════════
def parse_people(raw: bytes) -> list:
    df = pd.read_excel(io.BytesIO(raw), header=0)
    cols = [str(c).strip() for c in df.columns]

    # 성함 열 찾기: "성함"/"이름"/"name" 포함하는 열 우선, 없으면 B열(iloc[1])
    name_col = None
    for i, c in enumerate(cols):
        if any(k in c.lower() for k in ["성함", "이름", "name"]):
            name_col = i
            break
    if name_col is None:
        name_col = 1  # 기본값: B열

    # Q1~Q30 열 찾기: "q1"~"q30" 포함하는 열 우선, 없으면 name_col 다음 30개
    q_cols = {}
    for i, c in enumerate(cols):
        cl = c.lower().replace(" ", "").replace(".", "")
        for q in range(1, 31):
            if cl.startswith(f"q{q}") or cl == f"{q}":
                if q not in q_cols:
                    q_cols[q] = i
    if len(q_cols) < 30:
        # 열 이름으로 못 찾으면 위치 기반으로 fallback
        q_cols = {q: name_col + q for q in range(1, 31)}

    out = []
    for _, row in df.iterrows():
        name = str(row.iloc[name_col]).strip()
        if not name or name.lower() == "nan":
            continue
        scores = {}
        for q in range(1, 31):
            try:
                val = row.iloc[q_cols[q]]
                scores[str(q)] = float(val) if pd.notna(val) else 0.0
            except:
                scores[str(q)] = 0.0
        out.append({"name": name, "scores": scores})
    return out

# ══════════════════════════════════════════════════════════════════
# 엑셀 생성
# ══════════════════════════════════════════════════════════════════
def _copy_ws(wb, src, title):
    ws = wb.create_sheet(title=title)
    for cl, cd in src.column_dimensions.items():
        ws.column_dimensions[cl].width = cd.width
    for rn, rd in src.row_dimensions.items():
        ws.row_dimensions[rn].height = rd.height
    for row in src.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font=copy.copy(cell.font); nc.border=copy.copy(cell.border)
                nc.fill=copy.copy(cell.fill); nc.number_format=cell.number_format
                nc.protection=copy.copy(cell.protection); nc.alignment=copy.copy(cell.alignment)
    for m in src.merged_cells.ranges:
        ws.merge_cells(str(m))
    return ws

def build_excel(people, excel_tpl: bytes) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for p in people:
        src = load_workbook(io.BytesIO(excel_tpl)).worksheets[0]
        ws  = _copy_ws(wb, src, p["name"][:31])
        r   = compute(p["scores"])
        ws.cell(1,1).value = p["name"]
        for q in range(1, 31):
            ws.cell(q+3, 3).value = float(p["scores"].get(str(q), 0))
        for k, rl in COMPETENCY_MAP.items():
            avg = r["competency"][k]
            ws.cell(COMP_ROW[k],7).value = round(avg*len(rl),2); ws.cell(COMP_ROW[k],7).number_format="0.00"
            ws.cell(COMP_ROW[k],8).value = avg;                  ws.cell(COMP_ROW[k],8).number_format="0.00"
        for k, rl in SKILL_MAP.items():
            avg = r["skill_raw"][k]
            ws.cell(SKILL_ROW[k],7).value = round(avg*len(rl),2); ws.cell(SKILL_ROW[k],7).number_format="0.00"
            ws.cell(SKILL_ROW[k],8).value = avg;                   ws.cell(SKILL_ROW[k],8).number_format="0.00"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
# PPT 생성
# 베이스: 2슬라이드짜리 template_pptx.pptx
# slide1 → 1번째 사람, slide2 → 2번째 사람
# 3번째 이후 → slide2 구조 복제해서 추가
# ══════════════════════════════════════════════════════════════════
def _replace_chart_vals(chart_bytes, new_vals):
    s = chart_bytes.decode('utf-8')
    val_m = re.search(r'(<c:val>.*?<c:numCache>)(.*?)(</c:numCache>.*?</c:val>)', s, re.DOTALL)
    if not val_m: return chart_bytes
    before = re.sub(r'<c:ptCount val="\d+"/>', f'<c:ptCount val="{len(new_vals)}"/>', val_m.group(1))
    fmt = re.search(r'<c:formatCode>[^<]*</c:formatCode>', val_m.group(2))
    fmt_tag = fmt.group(0) if fmt else '<c:formatCode>0.00</c:formatCode>'
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i,v in enumerate(new_vals))
    s = (s[:val_m.start()] + before + f'{fmt_tag}<c:ptCount val="{len(new_vals)}"/>{pts}' + val_m.group(3) + s[val_m.end():])
    # y축(valAx) 범위 0~5 고정 - valAx 안의 scaling만 교체
    s = re.sub(
        r'(<c:valAx>.*?<c:scaling>).*?(</c:scaling>)',
        r'\1<c:orientation val="minMax"/><c:max val="5"/><c:min val="0"/>\2',
        s, flags=re.DOTALL
    )
    return s.encode('utf-8')

def _new_guids(s):
    for g in set(re.findall(r'\{[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}\}', s)):
        s = s.replace(g, '{'+str(uuid.uuid4()).upper()+'}')
    return s

def _ws_name(n): return f"Microsoft_Excel_Worksheet{n if n>0 else ''}.xlsx"

# ── 차트 색상 / 동그라미 위치 상수 (역산으로 구한 실제 플롯 영역) ──
_PHASE_PLOT_X = 2412488;  _PHASE_BAR_W = 1244294   # 6개 막대
_STRAT_PLOT_X = 2351917;  _STRAT_BAR_W = 763098    # 10개 막대
_CIRCLE_PHASE_Y  = 3213000;  _CIRCLE_PHASE_CY  = 437638
_CIRCLE_STRAT_Y  = 6021000;  _CIRCLE_STRAT_CY  = 465643

def _bar_cx_phase(idx):
    return int(_PHASE_PLOT_X + (idx + 0.5) * _PHASE_BAR_W)

def _bar_cx_strat(idx):
    return int(_STRAT_PLOT_X + (idx + 0.5) * _STRAT_BAR_W)

def _update_chart_phase_colors(chart_bytes, vals):
    """최대→파란색(4480B1), 최소→빨간색(C00000). 동값이면 모두 색칠."""
    s = chart_bytes.decode('utf-8')
    max_val = max(vals); min_val = min(vals)
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    dpts = ''
    # 최대값 인덱스 모두 파란색 (최대=최소인 경우 skip)
    if max_val != min_val:
        for idx in [i for i,v in enumerate(vals) if v == max_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="4480B1"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
        for idx in [i for i,v in enumerate(vals) if v == min_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="C00000"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _update_chart_strategy_colors(chart_bytes, vals):
    """전체 노란색(FFD000), 소프트평균(idx=3)·하드평균(idx=9)만 남색(2D5576)"""
    s = chart_bytes.decode('utf-8')
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    s = re.sub(
        r'(<c:spPr>)<a:solidFill>.*?</a:solidFill>',
        r'\1<a:solidFill><a:srgbClr val="FFD000"/></a:solidFill>',
        s, count=1, flags=re.DOTALL
    )
    dpts = ''
    for idx in [3, 9]:
        if idx < len(vals):
            dpts += (
                f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="2D5576"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>'
            )
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _move_circle(slide_str, circle_name, new_x, new_y, new_cx, new_cy):
    idx = slide_str.find(f'name="{circle_name}"')
    if idx == -1: return slide_str
    start = slide_str.rfind('<p:pic>', 0, idx)
    end   = slide_str.find('</p:pic>', idx) + len('</p:pic>')
    pic   = slide_str[start:end]
    pic   = re.sub(r'<a:off x="[^"]*" y="[^"]*"/>', f'<a:off x="{new_x}" y="{new_y}"/>', pic)
    pic   = re.sub(r'<a:ext cx="[^"]*" cy="[^"]*"/>', f'<a:ext cx="{new_cx}" cy="{new_cy}"/>', pic)
    return slide_str[:start] + pic + slide_str[end:]

def _get_strat_circle_targets(strat_vals):
    pull_vals = [(i, strat_vals[i]) for i in range(3)]
    push_vals = [(i, strat_vals[i]) for i in range(4, 9)]
    
    pull_max = max(v for _, v in pull_vals)
    push_max = max(v for _, v in push_vals)
    
    pull_targets = [i for i, v in pull_vals if v == pull_max]
    push_targets = [i for i, v in push_vals if v == push_max]
    
    # Pull: 3개 동률이면 0개 / Push: 3개 이상 동률이면 0개
    if len(pull_targets) == 3:
        pull_targets = []
    if len(push_targets) >= 3:
        push_targets = []
    
    return sorted(pull_targets + push_targets)

def _update_circles(slide_str, comp_vals, strat_vals):
    # phase: circle2=최대값 막대, circle1=최소값 막대
    max_val = max(comp_vals); min_val = min(comp_vals)
    max_idx = comp_vals.index(max_val)
    min_idx = comp_vals.index(min_val)
    cw_p = int(_PHASE_BAR_W * 0.85)
    slide_str = _move_circle(slide_str, 'circle2',
        _bar_cx_phase(max_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)
    slide_str = _move_circle(slide_str, 'circle1',
        _bar_cx_phase(min_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)

    # strategy: Pull/Push 각 평균 기준 ±0.3, 최대 3개
    targets = _get_strat_circle_targets(strat_vals)
    cw_s = int(_STRAT_BAR_W * 0.85)
    OFF_SCREEN = -2057475  # 슬라이드 밖 표들 사이 여백 (table_phase, table_strategy와 같은 영역)
    for ci in range(4):  # circle3, circle4, circle5, circle6
        if ci < len(targets):
            slide_str = _move_circle(slide_str, f'circle{ci+3}',
                _bar_cx_strat(targets[ci]) - cw_s//2, _CIRCLE_STRAT_Y, cw_s, _CIRCLE_STRAT_CY)
        else:
            # 사용 안 하는 circle은 화면 밖으로
            slide_str = _move_circle(slide_str, f'circle{ci+3}',
                OFF_SCREEN, 3450346, cw_s, _CIRCLE_STRAT_CY)  # 두 표 사이 y
    return slide_str

def _fill_slide(sl_str, person, result):
    c=result["competency"]; s=result["skill_raw"]; sa=result["soft_avg"]; ha=result["hard_avg"]
    comp_vals  = list(c.values())
    strat_vals = [s[k] for k in SOFT_SKILLS]+[sa]+[s[k] for k in HARD_SKILLS]+[ha]
    sl_str = sl_str.replace("{{NAME}}", person["name"])
    sl_str = _update_circles(sl_str, comp_vals, strat_vals)
    return sl_str

def build_ppt(people, ppt_tpl: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(ppt_tpl)) as src:
        infos = {info.filename: info for info in src.infolist()}
        files = {info.filename: src.read(info.filename) for info in src.infolist()}

    max_chart = max(int(m) for m in re.findall(r'chart(\d+)\.xml', ' '.join(files)))
    max_color = max(int(m) for m in re.findall(r'colors(\d+)\.xml', ' '.join(files)))
    max_style = max(int(m) for m in re.findall(r'style(\d+)\.xml', ' '.join(files)))
    ws_nums   = [int(m) if m else 0 for m in re.findall(r'Worksheet(\d*)\.xlsx', ' '.join(files))]
    max_ws    = max(ws_nums)

    prs_xml  = files["ppt/presentation.xml"]
    prs_rels = files["ppt/_rels/presentation.xml.rels"]
    ct_xml   = files["[Content_Types].xml"]
    max_sid  = max(int(m) for m in re.findall(r'<p:sldId id="(\d+)"', prs_xml.decode()))
    max_rid  = max(int(m) for m in re.findall(r'Id="rId(\d+)"', prs_rels.decode()))

    # slide2 기준 원본 (복제용)
    orig_s2  = files["ppt/slides/slide2.xml"]
    orig_s2r = files["ppt/slides/_rels/slide2.xml.rels"]
    orig_c3  = files["ppt/charts/chart3.xml"]
    orig_c4  = files["ppt/charts/chart4.xml"]
    orig_c3r = files["ppt/charts/_rels/chart3.xml.rels"]
    orig_c4r = files["ppt/charts/_rels/chart4.xml.rels"]

    for i, person in enumerate(people):
        result = compute(person["scores"])
        comp_vals  = list(result["competency"].values())
        strat_vals = ([result["skill_raw"][k] for k in SOFT_SKILLS] + [result["soft_avg"]] +
                      [result["skill_raw"][k] for k in HARD_SKILLS] + [result["hard_avg"]])

        if i == 0:
            sl = _fill_slide(files["ppt/slides/slide1.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide1.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart1.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart1.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart2.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart2.xml"], strat_vals), strat_vals)

        elif i == 1:
            sl = _fill_slide(files["ppt/slides/slide2.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide2.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart3.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart3.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart4.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart4.xml"], strat_vals), strat_vals)

        else:
            sn = i+1
            ca = max_chart+(i-1)*2+1; cb = ca+1
            cola = max_color+(i-1)*2+1; colb = cola+1
            sta  = max_style+(i-1)*2+1; stb  = sta+1
            wsa_n = max_ws+(i-1)*2+1;   wsb_n = wsa_n+1

            sl = _new_guids(orig_s2.decode('utf-8'))
            sl = _fill_slide(sl, person, result)
            files[f"ppt/slides/slide{sn}.xml"] = sl.encode('utf-8')
            files[f"ppt/slides/_rels/slide{sn}.xml.rels"] = (
                orig_s2r
                .replace(b"chart3.xml", f"chart{ca}.xml".encode())
                .replace(b"chart4.xml", f"chart{cb}.xml".encode())
            )
            files[f"ppt/charts/chart{ca}.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(orig_c3, comp_vals), comp_vals)
            files[f"ppt/charts/chart{cb}.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(orig_c4, strat_vals), strat_vals)
            files[f"ppt/charts/_rels/chart{ca}.xml.rels"] = (
                orig_c3r
                .replace(b"chart3.xml",  f"chart{ca}.xml".encode())
                .replace(b"colors3.xml", f"colors{cola}.xml".encode())
                .replace(b"style3.xml",  f"style{sta}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet2.xlsx", _ws_name(wsa_n).encode())
            )
            files[f"ppt/charts/_rels/chart{cb}.xml.rels"] = (
                orig_c4r
                .replace(b"chart4.xml",  f"chart{cb}.xml".encode())
                .replace(b"colors4.xml", f"colors{colb}.xml".encode())
                .replace(b"style4.xml",  f"style{stb}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet3.xlsx", _ws_name(wsb_n).encode())
            )
            files[f"ppt/charts/colors{cola}.xml"] = files["ppt/charts/colors3.xml"]
            files[f"ppt/charts/colors{colb}.xml"] = files["ppt/charts/colors4.xml"]
            files[f"ppt/charts/style{sta}.xml"]   = files["ppt/charts/style3.xml"]
            files[f"ppt/charts/style{stb}.xml"]   = files["ppt/charts/style4.xml"]
            files[f"ppt/embeddings/{_ws_name(wsa_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"]
            files[f"ppt/embeddings/{_ws_name(wsb_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"]

            def add_info(nn, rn):
                ni = zipfile.ZipInfo(nn); ni.compress_type = infos[rn].compress_type; infos[nn] = ni
            for nn, rn in [
                (f"ppt/slides/slide{sn}.xml",            "ppt/slides/slide2.xml"),
                (f"ppt/slides/_rels/slide{sn}.xml.rels", "ppt/slides/_rels/slide2.xml.rels"),
                (f"ppt/charts/chart{ca}.xml",            "ppt/charts/chart3.xml"),
                (f"ppt/charts/chart{cb}.xml",            "ppt/charts/chart4.xml"),
                (f"ppt/charts/_rels/chart{ca}.xml.rels", "ppt/charts/_rels/chart3.xml.rels"),
                (f"ppt/charts/_rels/chart{cb}.xml.rels", "ppt/charts/_rels/chart4.xml.rels"),
                (f"ppt/charts/colors{cola}.xml",         "ppt/charts/colors3.xml"),
                (f"ppt/charts/colors{colb}.xml",         "ppt/charts/colors4.xml"),
                (f"ppt/charts/style{sta}.xml",           "ppt/charts/style3.xml"),
                (f"ppt/charts/style{stb}.xml",           "ppt/charts/style4.xml"),
                (f"ppt/embeddings/{_ws_name(wsa_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"),
                (f"ppt/embeddings/{_ws_name(wsb_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"),
            ]: add_info(nn, rn)

            max_sid+=1; max_rid+=1; rid=f"rId{max_rid}"
            prs_xml  = prs_xml.replace(b'</p:sldIdLst>', f'<p:sldId id="{max_sid}" r:id="{rid}"/></p:sldIdLst>'.encode())
            prs_rels = prs_rels.replace(b'</Relationships>', f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{sn}.xml"/></Relationships>'.encode())
            ct_xml   = ct_xml.replace(b'</Types>',
                f'<Override PartName="/ppt/slides/slide{sn}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                f'<Override PartName="/ppt/charts/chart{ca}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'<Override PartName="/ppt/charts/chart{cb}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'</Types>'.encode())

    files["ppt/presentation.xml"]            = prs_xml
    files["ppt/_rels/presentation.xml.rels"] = prs_rels
    files["[Content_Types].xml"]             = ct_xml
    files["docProps/app.xml"] = re.sub(rb'<Slides>\d+</Slides>', f'<Slides>{len(people)}</Slides>'.encode(), files["docProps/app.xml"])

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w') as zout:
        for name, data in files.items():
            zout.writestr(infos[name], data)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════
# 템플릿 탐색
# ══════════════════════════════════════════════════════════════════
def find_template(ext: str):
    base = Path(__file__).parent
    found = sorted(base.glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    found = sorted(Path(os.getcwd()).glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    return None, None

# ══════════════════════════════════════════════════════════════════
# UI 스타일
# ══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="CLiCK | 리더십 영향력 진단",
    layout="wide",
    initial_sidebar_state="expanded"
)
 
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&family=DM+Sans:wght@300;400;500;600&display=swap');
 
html, body, [class*="css"] {
    font-family: 'Noto Sans KR', 'DM Sans', sans-serif;
}
 
/* 전체 배경 */
.stApp {
    background-color: #F5F6F8;
}
 
/* 사이드바 */
[data-testid="stSidebar"] {
    background-color: #0F2744;
    border-right: none;
}
[data-testid="stSidebar"] * {
    color: #CBD5E1 !important;
}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] strong {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stCaption {
    color: #64748B !important;
}
 
/* 메인 컨텐츠 패딩 */
.main .block-container {
    padding: 2rem 3rem 3rem 3rem;
    max-width: 900px;
}
 
/* 헤더 영역 */
.app-header {
    background: linear-gradient(135deg, #0F2744 0%, #1E3A5F 100%);
    border-radius: 12px;
    padding: 2.2rem 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: rgba(255,255,255,0.04);
    border-radius: 50%;
}
.app-header::after {
    content: '';
    position: absolute;
    bottom: -60px; right: 60px;
    width: 140px; height: 140px;
    background: rgba(255,255,255,0.03);
    border-radius: 50%;
}
.app-header .badge {
    display: inline-block;
    background: rgba(255,255,255,0.12);
    color: #93C5FD;
    font-size: 0.72rem;
    font-weight: 500;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    padding: 0.3rem 0.8rem;
    border-radius: 20px;
    margin-bottom: 0.9rem;
}
.app-header h1 {
    color: #FFFFFF;
    font-size: 1.65rem;
    font-weight: 700;
    margin: 0 0 0.4rem 0;
    letter-spacing: -0.01em;
    line-height: 1.3;
}
.app-header p {
    color: #93AECF;
    font-size: 0.88rem;
    margin: 0;
    font-weight: 400;
}
 
/* 안내 카드 */
.guide-card {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-left: 4px solid #1E3A5F;
    border-radius: 10px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1.6rem;
}
.guide-card .guide-title {
    font-size: 0.8rem;
    font-weight: 600;
    color: #1E3A5F;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 0.7rem;
}
.guide-card ul {
    margin: 0;
    padding-left: 1.2rem;
    color: #475569;
    font-size: 0.875rem;
    line-height: 1.8;
}
.guide-card ul li::marker {
    color: #1E3A5F;
}
 
/* 업로드 섹션 라벨 */
.section-label {
    font-size: 0.78rem;
    font-weight: 600;
    color: #64748B;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
}
 
/* 파일 업로더 커스텀 */
[data-testid="stFileUploader"] {
    background: #FFFFFF;
    border: 1.5px dashed #CBD5E1;
    border-radius: 10px;
    padding: 0.5rem;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #1E3A5F;
}
 
/* 생성 버튼 */
.stButton > button {
    background: linear-gradient(135deg, #0F2744, #1E3A5F);
    color: #FFFFFF;
    border: none;
    border-radius: 8px;
    font-family: 'Noto Sans KR', sans-serif;
    font-size: 0.92rem;
    font-weight: 600;
    padding: 0.7rem 1.5rem;
    letter-spacing: 0.02em;
    width: 100%;
    transition: all 0.2s ease;
    box-shadow: 0 2px 8px rgba(15,39,68,0.25);
}
.stButton > button:hover {
    background: linear-gradient(135deg, #1a3a5c, #265080);
    box-shadow: 0 4px 14px rgba(15,39,68,0.35);
    transform: translateY(-1px);
}
.stButton > button:active {
    transform: translateY(0);
}
 
/* 성공 메시지 */
.result-banner {
    background: linear-gradient(135deg, #ECFDF5, #D1FAE5);
    border: 1px solid #6EE7B7;
    border-radius: 10px;
    padding: 1.1rem 1.5rem;
    margin: 1.2rem 0;
    display: flex;
    align-items: center;
    gap: 0.7rem;
}
.result-banner .icon { font-size: 1.3rem; }
.result-banner .text {
    color: #065F46;
    font-size: 0.9rem;
    font-weight: 500;
}
 
/* 다운로드 버튼 그룹 */
.download-section {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 1.4rem 1.6rem;
    margin-top: 0.8rem;
}
.download-section .dl-title {
    font-size: 0.78rem;
    font-weight: 600;
    color: #94A3B8;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    margin-bottom: 1rem;
}
[data-testid="stDownloadButton"] > button {
    background: #F8FAFC;
    color: #1E3A5F;
    border: 1.5px solid #CBD5E1;
    border-radius: 8px;
    font-family: 'Noto Sans KR', sans-serif;
    font-size: 0.85rem;
    font-weight: 500;
    padding: 0.55rem 1rem;
    width: 100%;
    transition: all 0.18s ease;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #EFF6FF;
    border-color: #1E3A5F;
    color: #0F2744;
}
 
/* 구분선 */
hr {
    border: none;
    border-top: 1px solid #E2E8F0;
    margin: 1.5rem 0;
}
 
/* 에러 메시지 */
.stAlert {
    border-radius: 8px;
    font-size: 0.875rem;
}
</style>
""", unsafe_allow_html=True)
 
# ── 사이드바 ──
with st.sidebar:
    st.markdown("### 📋 문항 매핑 참고")
    st.caption("진단 문항과 역량 매핑 기준표")
 
    def find_image(filename):
        base = Path(__file__).parent
        p = base / filename
        if p.exists(): return str(p)
        p2 = Path(os.getcwd()) / filename
        if p2.exists(): return str(p2)
        return None
 
    img_stage    = find_image("mapping_stage.jpg")
    img_strategy = find_image("mapping_strategy.jpg")
 
    if img_stage:
        st.markdown("**▶ 리더십 영향력 단계**")
        st.image(str(img_stage), use_container_width=True)
    else:
        st.info("mapping_stage.jpg를 루트에 업로드해주세요")
 
    if img_strategy:
        st.markdown("**▶ 리더십 영향력 전략**")
        st.image(str(img_strategy), use_container_width=True)
    else:
        st.info("mapping_strategy.jpg를 루트에 업로드해주세요")
 
    st.markdown("---")
    st.markdown("**CLiCK**")
    st.caption("리더십 영향력 진단 자동화 시스템 v1.0")
 
# ── 메인 헤더 ──
st.markdown("""
<div class="app-header">
    <div class="badge">CLiCK</div>
    <h1>리더십 영향력 진단 결과 자동화 (구글 폼 응답용)</h1>
    <p>구글 폼 응답 데이터를 업로드하면 개인별 진단 보고서(Excel · PPT)를 자동으로 생성합니다.</p>
</div>
""", unsafe_allow_html=True)
 
# ── 안내 카드 ──
st.markdown("""
<div class="guide-card">
    <div class="guide-title">📌 업로드 전 확인사항</div>
    <ul>
        <li>구글 폼 응답을 <strong>엑셀(.xlsx)로 다운로드</strong>한 파일을 업로드하세요.</li>
        <li>파일 형식: <strong>1행 = 헤더, 2행부터 = 응답자 데이터</strong> (구글 폼 기본 형식)</li>
        <li>응답자별 개인 엑셀 파일이 있다면 → <a href="https://leadershipinfluencev2-htierxuqwg8zwdv3afcqbr.streamlit.app/" target="_blank">개인 엑셀 파일용 자동화</a>를 이용하세요.</li>
    </ul>
</div>
""", unsafe_allow_html=True)
 
# ── 파일 업로더 ──
st.markdown('<div class="section-label">응답 데이터 업로드</div>', unsafe_allow_html=True)
response_file = st.file_uploader(
    "구글 폼 응답 엑셀 파일을 선택하세요",
    type=["xlsx", "xls"],
    label_visibility="collapsed"
)
 
st.markdown("<hr>", unsafe_allow_html=True)
 
# ── 생성 버튼 ──
if st.button("보고서 생성", type="primary", use_container_width=True):
    if response_file is None:
        st.error("응답 엑셀 파일을 먼저 업로드해주세요.")
        st.stop()
 
    resp_bytes = response_file.read()
 
    excel_tpl, ep = find_template(".xlsx")
    if not excel_tpl:
        st.error("엑셀 템플릿 파일을 찾을 수 없습니다. GitHub 루트에 .xlsx 파일을 업로드해주세요.")
        st.stop()
 
    ppt_tpl, pp = find_template(".pptx")
    if not ppt_tpl:
        st.error("PPT 템플릿 파일을 찾을 수 없습니다. GitHub 루트에 .pptx 파일을 업로드해주세요.")
        st.stop()
 
    try:
        people = parse_people(resp_bytes)
    except Exception as e:
        st.error(f"파일 파싱 중 오류가 발생했습니다: {e}")
        st.code(traceback.format_exc())
        st.stop()
 
    if not people:
        st.error("응답자 데이터를 찾을 수 없습니다. 파일 형식을 확인해주세요.")
        st.stop()
 
    with st.spinner(f"{len(people)}명의 보고서를 생성하고 있습니다..."):
        try:
            excel_out = build_excel(people, excel_tpl)
        except Exception as e:
            st.error(f"엑셀 생성 중 오류가 발생했습니다: {e}")
            st.code(traceback.format_exc())
            st.stop()
        try:
            ppt_out = build_ppt(people, ppt_tpl)
        except Exception as e:
            st.error(f"PPT 생성 중 오류가 발생했습니다: {e}")
            st.code(traceback.format_exc())
            st.stop()
 
    st.session_state["excel_out"] = excel_out
    st.session_state["ppt_out"]   = ppt_out
    st.session_state["n"]         = len(people)
    st.session_state["done"]      = True
 
# ── 완료 및 다운로드 ──
if st.session_state.get("done"):
    excel_out = st.session_state["excel_out"]
    ppt_out   = st.session_state["ppt_out"]
    n         = st.session_state["n"]
 
    st.markdown(f"""
    <div class="result-banner">
        <span class="icon">✅</span>
        <span class="text">보고서 생성 완료 — 총 <strong>{n}명</strong> | Excel {n}시트 · PPT {n}슬라이드</span>
    </div>
    """, unsafe_allow_html=True)
 
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("CIAM_리더십영향력_진단지.xlsx", excel_out)
        zf.writestr("CIAM_리더십영향력_진단결과.pptx", ppt_out)
 
    st.markdown('<div class="download-section"><div class="dl-title">📥 파일 다운로드</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button(
            "전체 ZIP 다운로드",
            data=zip_buf.getvalue(),
            file_name="CIAM_리더십영향력_결과.zip",
            mime="application/zip",
            use_container_width=True
        )
    with col2:
        st.download_button(
            "Excel만 다운로드",
            data=excel_out,
            file_name="CIAM_리더십영향력_진단지.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col3:
        st.download_button(
            "PPT만 다운로드",
            data=ppt_out,
            file_name="CIAM_리더십영향력_진단결과.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True
        )
    st.markdown('</div>', unsafe_allow_html=True)
 
